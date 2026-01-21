# =============================================================================
# HERMES PLATFORM - Export Router
# =============================================================================
# Excel export endpoint'leri (FR 4.x). Sadece Admin erişebilir.
# =============================================================================

from typing import Optional
from datetime import date
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from shared.auth import require_admin, CurrentUser
from ..services.report_service import ReportService


router = APIRouter(prefix="/export", tags=["Export"])


@router.get(
    "/excel/v1",
    summary="Excel Export (v1)",
    description="""
    Zaman girişlerini Excel dosyası olarak dışa aktarır (FR 4.3).
    
    Çıktı formatı:
    - Tarih
    - Müşteri
    - Proje
    - İş Tipi
    - Süre (Saat)
    - Kişi
    - Açıklama
    """
)
async def export_excel_v1(
    request: Request,
    start_date: Optional[date] = Query(None, description="Başlangıç tarihi"),
    end_date: Optional[date] = Query(None, description="Bitiş tarihi"),
    admin: CurrentUser = Depends(require_admin)
):
    """
    Excel export endpoint'i.
    
    Varsayılan tarih aralığı: Son 30 gün.
    """
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.replace("Bearer ", "")
    
    try:
        service = ReportService(token)
        excel_data = await service.get_excel_data(start_date, end_date)
        
        # Excel dosyası oluştur
        excel_buffer = create_excel_file(excel_data)
        
        # Dosya adı oluştur
        start_str = (start_date or date.today()).isoformat()
        end_str = (end_date or date.today()).isoformat()
        filename = f"hermes_rapor_{start_str}_{end_str}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Excel raporu oluşturulamadı: {str(e)}"
        )


def create_excel_file(data: list) -> BytesIO:
    """
    Veriyi Excel dosyasına dönüştürür.
    
    openpyxl kütüphanesi kullanılır.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # openpyxl yoksa basit CSV döndür
        return create_csv_fallback(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Zaman Girişleri"
    
    # Başlık stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = ["Tarih", "Müşteri", "Proje", "İş Tipi", "Süre (Saat)", "Kişi", "Açıklama"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Veriler
    for row_idx, row_data in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row_data.get("Tarih", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=row_data.get("Müşteri", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=row_data.get("Proje", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=row_data.get("İş Tipi", "")).border = thin_border
        ws.cell(row=row_idx, column=5, value=row_data.get("Süre (Saat)", 0)).border = thin_border
        ws.cell(row=row_idx, column=6, value=row_data.get("Kişi", "")).border = thin_border
        ws.cell(row=row_idx, column=7, value=row_data.get("Açıklama", "")).border = thin_border
    
    # Kolon genişlikleri
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 40
    
    # Buffer'a kaydet
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_csv_fallback(data: list) -> BytesIO:
    """openpyxl yoksa CSV formatında döndür."""
    import csv
    
    buffer = BytesIO()
    # UTF-8 BOM ekle (Excel için)
    buffer.write(b'\xef\xbb\xbf')
    
    text_buffer = BytesIO()
    writer = csv.DictWriter(
        text_buffer,
        fieldnames=["Tarih", "Müşteri", "Proje", "İş Tipi", "Süre (Saat)", "Kişi", "Açıklama"]
    )
    writer.writeheader()
    writer.writerows(data)
    
    buffer.write(text_buffer.getvalue())
    buffer.seek(0)
    return buffer
